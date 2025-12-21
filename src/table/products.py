# ----------------------------------------------------------------------------#
# Embedded libraries                                                          #
# ----------------------------------------------------------------------------#
from os.path import isfile

# ----------------------------------------------------------------------------#
# Project modules                                                             #
# ----------------------------------------------------------------------------#
from src.utilities.utilities import creating_necessary_folders

# ----------------------------------------------------------------------------#
# External libraries                                                          #
# ----------------------------------------------------------------------------#
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from pandas import DataFrame, Series, ExcelWriter as pd_ExcelWriter, concat as pd_concat


def update_length_column_names(df_products:DataFrame, file_path:str, sheet_name:str='Quest 1') -> None:
    if not isfile(path=file_path):
        workbook = Workbook()
        workbook.active.title = 'Quest 1'
        workbook.save(filename=file_path)
    with pd_ExcelWriter(path=file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_products.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df_products.columns):
            max_length = len(col)
            column_letter = get_column_letter(col_idx=idx + 1)
            worksheet.column_dimensions[column_letter].width = max_length + 2


def create_table(list_products:dict, table_name:str='WB products', sheet_name:str='Quest 1') -> None:
    col_names = [
        'Ссылка на товар',
        'Артикул',
        'Название',
        'Цена',
        'Цена со скидкой',
        'Название селлера',
        'Ссылка на селлера',
        'Размеры товара',
        'Остатки по товару',
        'Рейтинг',
        'Количество отзывов',
        'Описание',
        'Ссылки на изображения',
        'Все характеристики с сохранением их структуры'
        ]
    list_series = list()
    for product in list_products:
        list_product = list()
        list_product.append(product['link'])
        list_product.append(product['article'])
        list_product.append(product['name'])
        list_product.append(product['fullPrice'])
        list_product.append(product['discountedPrice'])
        list_product.append(product['nameSeller'])
        list_product.append(product['linkSeller'])
        str_sizes = ""
        for number_size in range(len(product['productSizes'])):
            str_sizes += f'название: {product['productSizes'][number_size]['name']}, цена: {product['productSizes'][number_size]['fullPrice']}, цена по скидке: {product['productSizes'][number_size]['discountedPrice']};\n'
        list_product.append(str_sizes[:-1])
        list_product.append(product['productBalances'])
        list_product.append(product['reviewRating'])
        list_product.append(product['quantityFeedbacks'])
        list_product.append(product['description'])
        list_product.append(product['images'])
        str_characteristics = ""
        for characteristic in product['characteristics']:
            name, value = characteristic.popitem()
            str_characteristics += f'{name}: {value}\n'
        list_product.append(str_characteristics[:-1])
        row_series = Series(data=list_product, index=col_names)
        list_series.append(row_series)
    df_products = pd_concat(objs=list_series, axis=1)
    df_products = df_products.T
    creating_necessary_folders('docs')
    update_length_column_names(df_products=df_products, file_path=f"docs/{table_name}.xlsx", sheet_name=sheet_name)

